from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import pickle
import csv
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
from email.utils import parsedate_to_datetime
import pandas as pd
from datetime import datetime, timezone
import base64

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/spreadsheets"
]

CREDENTIALS_PATH = "/etc/secrets/credentials.json"
TOKEN_PATH_B64 = "/etc/secrets/token.pickleb64"

EXCEL_PATH = "result/log2.xlsx"

MESSAGE_ID = "Message_ID"

def authenticate():
    flow = InstalledAppFlow.from_client_secrets_file(
        'credentials.json', SCOPES)
    creds = flow.run_local_server(port=0)
    # Save token for reuse
    with open('token.pickle', 'wb') as token:
        pickle.dump(creds, token)
    return creds

import pickle

def load_credentials():
    if os.path.exists("/etc/secrets/credentials.json"):
        print("✅ Using Render credentials...")
        # Load base64-encoded pickle
        with open("/etc/secrets/token.pickleb64", "rb") as f:
            b64_data = f.read()
        raw_data = base64.b64decode(b64_data)
        creds = pickle.loads(raw_data)
        return creds
    else:
        print("✅ Using local credentials...")
        # Local development credentials
        with open("token.pickle", "rb") as token:
            creds = pickle.load(token)
        return creds

def load_render_credentials():
    try:
        # Load base64-encoded pickle
        with open(TOKEN_PATH_B64, "rb") as f:
            b64_data = f.read()
        raw_data = base64.b64decode(b64_data)
        creds = pickle.loads(raw_data)
        return creds

    except Exception as e:
        print("❌ Failed to load token.pickle. Re-authentication required.")
        print(str(e))
        # Optionally re-authenticate here if you want (requires user intervention)
        flow = InstalledAppFlow.from_client_secrets_file(
            CREDENTIALS_PATH, SCOPES
        )
        creds = flow.run_console()  # Only works if you can paste the auth code
        return creds

def get_all_unread_emails(service):
    # Retrieve unread messages
    results = service.users().messages().list(
        userId='me',
        labelIds=['INBOX'],
        q='is:unread'
    ).execute()
    messages = results.get('messages', [])
    question_emails = []

    for msg in messages:
        message_id = msg['id']

        message = service.users().messages().get(
            userId='me',
            id=message_id,
            format='metadata',
            metadataHeaders=['From', 'Subject', 'Date']
        ).execute()

        headers = {h['name']: h['value'] for h in message['payload']['headers']}
        
        question_emails.append(headers)

    return question_emails

def get_unread_question_emails(service, excel_path):
    # Load existing IDs from Excel
    try:
        df = pd.read_excel(excel_path)
        logged_ids = set(df[MESSAGE_ID].astype(str))
    except FileNotFoundError:
        logged_ids = set()
    
    # Retrieve unread messages
    results = service.users().messages().list(
        userId='me',
        labelIds=['INBOX'],
        q='is:unread'
    ).execute()
    messages = results.get('messages', [])
    question_emails = []

    for msg in messages:
        message_id = msg['id']

        # Skip if already logged
        if message_id in logged_ids:
            continue

        message = service.users().messages().get(
            userId='me',
            id=message_id,
            format='metadata',
            metadataHeaders=['From', 'Subject', 'Date']
        ).execute()

        headers = {h['name']: h['value'] for h in message['payload']['headers']}
        
        if '@Question' in headers.get('Subject', ''):
            # Attach message ID so you can save it later
            headers[MESSAGE_ID] = message_id
            question_emails.append(headers)

    return question_emails


def append_to_sheet(sheet_service, spreadsheet_id, data):
    body = {
        'values': [data]
    }
    sheet_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range='Sheet1!A:E',
        valueInputOption='USER_ENTERED',
        body=body
    ).execute()



def append_to_csv(filename, records):
        """
        records: List of lists, e.g.,
        [
            ["someone@gmail.com", "2025-07-01", "12:00:00", "0:15:00", "Not started"],
            ...
        ]
        """
        # If file doesn't exist yet, write header
        try:
            with open(filename, 'x', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["Email Sender", "Date", "Time Received", "Wait Time", "Status"])
                for row in records:
                    writer.writerow(row)
        except FileExistsError:
            # File exists, just append rows
            with open(filename, 'a', newline='') as csvfile:
                writer = csv.writer(csvfile)
                for row in records:
                    writer.writerow(row)




def append_to_excel(filename, records):
    if not os.path.exists(filename):
        # Create workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Log"

        headers = [MESSAGE_ID, "Email Sender", "Date", "Time Received", "Wait Time", "Status"]
        ws.append(headers)

        # Add dropdown validation to Status column
        dv = DataValidation(type="list", formula1='"Not started, In Progress, Done"', allow_blank=True)
        ws.add_data_validation(dv)

        # Apply validation to all rows in Status column
        max_row = 1048576  # Excel max rows
        status_col = get_column_letter(headers.index("Status")+1)  # E
        dv.add(f"{status_col}2:{status_col}{max_row}")

        wb.save(filename)
        print(f"✅ Created {filename} with headers and dropdowns.")

    # Load the workbook to append
    wb = load_workbook(filename)
    ws = wb.active

    # Append each record
    for row in records:
        ws.append(row)

    wb.save(filename)
    print(f"✅ Appended {len(records)} record(s) to {filename}.")

def format_wait_time(delta):
    """
    delta: datetime.timedelta
    Returns a string like "2 days, 3 hours, 15 minutes"
    """
    total_seconds = int(delta.total_seconds())
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    return f"{days} days, {hours} hours, {minutes} minutes"



def main():


    creds = load_credentials()
    
    gmail_service = build('gmail', 'v1', credentials=creds)

    emails = get_unread_question_emails(gmail_service, EXCEL_PATH)

    for email in emails:
        sender = email.get('From')
        date_str = email.get('Date')

        # Parse date correctly
        email_date = parsedate_to_datetime(date_str)
        if email_date.tzinfo is None:
            email_date = email_date.replace(tzinfo=timezone.utc)
        email_date = email_date.astimezone(timezone.utc)

        now = datetime.now(timezone.utc)
        delta = now - email_date
        wait_time = format_wait_time(delta)

        date_received = email_date.strftime('%Y-%m-%d')
        time_received = email_date.strftime('%H:%M:%S')

        message_id = email[MESSAGE_ID] 

        row = [
            message_id,
            sender,
            date_received,
            time_received,
            wait_time,
            "Not started"
        ]

        # check to see if result folder exists
        if not os.path.exists("result"):
            os.makedirs("result")

        append_to_excel(EXCEL_PATH, [row])

  

if __name__ == '__main__':
    main()